# coding=utf-8
import time
import csv
import io
import urllib
import requests
import sys
import urllib.request
import tkinter
import datetime
import time 
import threading
import json
import re
import codecs
import base64
import openpyxl 
from openpyxl import Workbook
from shutil import copyfileobj
from bs4 import BeautifulSoup
from yahoo_finance import Share
from pip._vendor.distlib.compat import raw_input
from tkinter import Label, Entry, Frame,SE,Tk, StringVar, Event
from tkinter.constants import INSERT,END
from tkinter.ttk import *
from time import sleep
from threading import Thread
import _thread
from asyncio import events
class InTimeData():
    def __init__(self):
        self.IntimeRquest = requests.session()
        self.IntimeRquest.get('http://mis.twse.com.tw/stock/index.jsp',headers = {'Accept-Language':'zh-TW,zh;q=0.8,en-US;q=0.6,en;q=0.4'})
        self.IntimeStock=[]
        self.InTimeFuture=[]
        self.PowerStage=0
        self.WeakStage=0
        self.MiddleStage=0
        self.FutureNowPrice=0
                
    def GetInTimeStockInfo(self,StockNumber):
        url = "http://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch=tse_"+StockNumber.__str__()+".tw"
        IntimeResponse = self.IntimeRquest.get(url)
        StockInfo = json.loads(IntimeResponse.text)
        self.IntimeStock.append(StockInfo)
    #    print(StockInfo['msgArray'][0]['v'])
        return StockInfo['msgArray'][0]['v']
    
    def FutureInTime(self):        #Get Future Price In time 
        url="http://mis.twse.com.tw/stock/data/futures_side.txt"   
        FutureIntTimeResponse = requests.get(url)
        FutureInfo=json.loads(FutureIntTimeResponse.text)
        self.FutureNowPrice =float(FutureInfo['msgArray'][0]['z'])
        self.MiddleStage = int((float(FutureInfo['msgArray'][0]['h'])+float(FutureInfo['msgArray'][0]['l']))/2)
        Stage=int(float(FutureInfo['msgArray'][0]['h'])-float(FutureInfo['msgArray'][0]['l']))
        self.PowerStage = int(Stage*0.618)
        self.WeakStage = int(Stage*0.382)
        

class GetData():
    def __init__(self):
        self.Responese = urllib.request.urlopen('http://www.twse.com.tw/ch/trading/exchange/BWIBBU/BWIBBU_d.php')
        self.soup = BeautifulSoup(self.Responese,"html.parser")
        self.TargetPERatio = 0  # raw_input("Please input the PERatio")
        self.TargetEPS = 0      #raw_input("Please input the EPS")
        self.TargetVolume=0
        self.TargetPriceLow=1
        self.TargetPrice=70
        self.AllData=[]
        self.TrendThreeMen=[]
        self.HistoryData=[]
        self.PERatioCsvfile()
#####################
    
    
    def PERatioCsvfile(self):
        first = requests.get('http://www.twse.com.tw/ch/trading/exchange/BWIBBU/BWIBBU_d.php')
        first.encoding = 'big5'
        soup = BeautifulSoup(first.text,"html.parser")
        a={}
        for hey in soup.select('input'):
            if 'hidden' in hey.get('type'):
                if hey.get('name') =='html':
                    a[hey.get('name')] = base64.b64encode(hey.get('value').encode('utf-8'))
                else:
                    a[hey.get('name')] = hey.get('value')
        data=a
        res2 = requests.post('http://www.twse.com.tw/ch/trading/exchange/BWIBBU/BWIBBU_print.php?language=ch&save=csv',data=a,stream=True)
        f = open('export.csv','wb')
        copyfileobj(res2.raw,f)
        f.close()
        self.SelectStock()
    def SelectStock(self):
        wb = Workbook()
        ws = wb.active
        with open('export.csv',newline='') as f:
            reader = csv.reader(f)
            index = 2
            title = ["代號","名稱","股價","本益比","EPS"]
            for x in range(0,5,1):
                ws.cell(row=1,column=x+1,value=title[x])
            for row in reader:
                if row!=[]:
                    try:
                        if int(row[0]) < 10000:
                        #    print(row[2])
                            if "-" not in (row[2]) and float(row[2]) < 20:
                             #   print(row)
                                Query = Share(row[0]+'.TW')
                                #print(float(Query.get_earnings_share()))     
                                #print(float(Query.get_open()))
                                #print("******************")
                                if float(Query.get_open()) >= 20 and float(Query.get_open()) <= 90 and float(Query.get_earnings_share())>1.2:
                                    print(row)
                                    ws.cell(row=index , column=1 , value = row[0])
                                    ws.cell(row=index , column=2 , value = row[1])
                                    ws.cell(row=index , column=3 , value = Query.get_open())
                                    ws.cell(row=index , column=4 , value = row[2])
                                    ws.cell(row=index , column=5 , value = Query.get_earnings_share())
                                    index = index + 1
                
                                    
                    except:
                        pass    
            wb.save("Final.xlsx")        
                    
    def SearchPERatio(self):
        for x in range(8,1503,5):
       # for x in range(8,self.soup.select('.basic2').__len__(),5):
            if self.soup.select('.basic2')[x].text != '-':
                if float(self.soup.select('.basic2')[x].text) <=float(self.TargetPERatio) :
                    QueryPrice = Share(self.soup.select('.basic2')[x-2].text+'.TW') #Query the API  
                    if QueryPrice.get_earnings_share()!=None and float(QueryPrice.get_earnings_share())>=self.TargetEPS:                     # Get EPS
                        self.AllData.append(self.soup.select('.basic2')[x-2].text)  # StockNumber
                        self.AllData.append(QueryPrice.get_open())                  # Get FinalPrice
                        self.AllData.append(self.soup.select('.basic2')[x].text)    # Get PERatio
                        self.AllData.append(QueryPrice.get_earnings_share())
    
    def DayMoving(self,StockNumber):        #個股的平均線
        PrehistoryURL = "http://real-chart.finance.yahoo.com/table.csv?s="+StockNumber+".TW"
        CSVReader = csv.reader(io.TextIOWrapper(urllib.request.urlopen(PrehistoryURL)))
        self.HistoryData.append(StockNumber)
        Day=0
        MovingPrice=0
        for row in CSVReader:            #Moving Days 
            if row[6]=='Adj Close':
                print(row)
                continue
            Day = Day+1
            MovingPrice = float(row[6])+MovingPrice
            if Day==5:
                self.HistoryData.append(float(MovingPrice/5))
            elif Day==15:
                self.HistoryData.append(float(MovingPrice/15))
            elif Day==30:
                self.HistoryData.append(float(MovingPrice/30))
                break
        return self.HistoryData
     
    def ThreeMenDollar(self): #三大法人
###############弱五點前則抓前一天資料##################
        Today = datetime.datetime.now().strftime("%Y%m%d")
        year =  int(time.strftime("%Y"))
        month= int(time.strftime("%m"))
        day = int(time.strftime("%d"))
        hour = float(time.strftime("%H"))
        judgeday = datetime.date(year,month,day).isoweekday()
        if judgeday==6:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
            Today = yesterday.strftime("%Y%m%d")
        elif judgeday==7:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=2)
            Today = yesterday.strftime("%Y%m%d")   
        elif judgeday==1 and hour < 17 :
            yesterday = datetime.datetime.now() - datetime.timedelta(days=3)
            Today = yesterday.strftime("%Y%m%d")


        print(Today)
        url = "http://www.twse.com.tw/ch/trading/fund/BFI82U/BFI82U_print.php?begin_date="+Today+"&report_type=day&language=ch&save=csv" 
        self.ThreeMenDollarResponse = urllib.request.urlopen(url)

    def FutureToday(self): #期貨 http://www.taifex.com.tw/chinese/3/7_12_8.asp
        
        TodayFuture={  #setting the post query data
              'syear':datetime.datetime.now().strftime("%Y"),
              'smonth':datetime.datetime.now().strftime("%m"),
              'sday':datetime.datetime.now().strftime("%d"),
              'eyear':datetime.datetime.now().strftime("%Y"),
              'emonth':datetime.datetime.now().strftime("%m"),
              'eday':datetime.datetime.now().strftime("%d"),
        }
##############如果是再下午五點前，則前一天資料#####################
        Today = datetime.datetime.now().strftime("%Y%m%d")
        year =  int(time.strftime("%Y"))
        month= int(time.strftime("%m"))
        day = int(time.strftime("%d"))
        hour = float(time.strftime("%H"))
        judgeday = datetime.date(year,month,day).isoweekday()
        if judgeday==6:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
            TodayFuture['sday'] = yesterday.strftime("%d")
            TodayFuture['eday'] = yesterday.strftime("%d")
        elif judgeday==7:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=2)
            TodayFuture['sday'] = yesterday.strftime("%d")
            TodayFuture['eday'] = yesterday.strftime("%d")
        elif judgeday==1 and hour < 17 :
            yesterday = datetime.datetime.now() - datetime.timedelta(days=3)
            TodayFuture['sday'] = yesterday.strftime("%d")
            TodayFuture['eday'] = yesterday.strftime("%d")
            
        elif hour < 17:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
            TodayFuture['sday'] = yesterday.strftime("%d")
            TodayFuture['eday'] = yesterday.strftime("%d")
            
            
        FutureRequest= requests.post('http://www.taifex.com.tw/chinese/3/7_12_8dl.asp',data=TodayFuture,allow_redirects=False)
        URLRedirect = "http://www.taifex.com.tw"+FutureRequest.headers['Location']
        RealCSVFile = requests.get(URLRedirect)
        print(RealCSVFile.url)
        self.TodayFutureResponse = urllib.request.urlopen(RealCSVFile.url)

    


class GUI(GetData):
    
    def __init__(self):
        super().__init__()
####################### initial interface #######################
        
        self.LabelName=['滴滿足點','弱關','中關','強關','高滿足點']
        self.BigThreeLabelName=['法人','自營商','投信','外資','三大法人']
        self.InTimeObject = InTimeData()
        self.InTimeObject.FutureInTime()
               
        self.interface = tkinter.Tk()
        self.interface.title("HI")
              
        self.aa=tkinter.Scrollbar()
        self.ThreemenBox=tkinter.Listbox(self.interface)
        self.ThreemenBox.grid(row=6,column=4)
        self.aa.config(command=self.ThreemenBox.yview())
        self.ThreemenBox.config(yscrollcommand=self.aa)

###############################################################################        
        self.inputPERatio = Label()
        self.inputPERatio["text"]="Input Ratio"
        self.inputPERatio.grid(row=0,column=0)
        self.inputPREField = Entry(self.interface)
        self.inputPREField.bind('<Return>',self.GetTextFromPREField)
        self.inputPREField.grid(row=0,column=1)
        
        self.inputEPS = Label()
        self.inputEPS["text"]="Input EPS"
        self.inputEPS.grid(row=1,column=0)
        self.inputEPSField = Entry(self.interface)
        self.inputEPSField.bind('<Return>',self.GetTextFromEPSField)
        self.inputEPSField.grid(row=1,column=1)
        
        self.inputVolume = Label()
        self.inputVolume["text"]="Input Volume"
        self.inputVolume.grid(row=2,column=0)
        self.inputVolumeField = Entry(self.interface)
        self.inputVolumeField.bind('<Return>',self.GetTextFromVolumeField)
        self.inputVolumeField.grid(row=2,column=1)
        
        self.inputWannaPrice = Label()
        self.inputWannaPrice["text"]="Input Close Price"
        self.inputWannaPrice.grid(row=3,column=0)
        self.inputWannaPriceField = Entry(self.interface)
        self.inputWannaPriceField.bind('<Return>',self.GetTextFromWannaPriceField)
        self.inputWannaPriceField.grid(row=3,column=1)
         
        self.CheckButton = tkinter.Button(text="Enter")
        self.CheckButton.bind('<Button-1>',self.tesfile)
        self.CheckButton.grid(row=4)
###############################################################################    
#期貨未平倉口數Label設定

        for x in range(0,5,1):
            self.InTimeFutureLabel = Label()  
            self.InTimeFutureLabel["text"]=self.BigThreeLabelName[x]
            self.InTimeFutureLabel.grid(row=x,column=4)

        self.FutureNameTitle=Label()
        self.FutureNameTitle["text"]="未平倉口數"
        self.FutureNameTitle.grid(row=0,column=5)
        self.FutureNameTitle=Label()
        self.FutureNameTitle["text"]="買賣超金額"
        self.FutureNameTitle.grid(row=0,column=6)
  
        self.HighSatisfy = StringVar()
        self.showWeakStage = StringVar()
        self.showPowerStage = StringVar()
        self.showMiddleStage=StringVar()
        self.LowSatisfy = StringVar()
        self.showClosePrice = StringVar()
        
        self.LabelLowSatisfy=Label()
        self.LabelLowSatisfy.grid(row=1,column=8)
        self.LabelWeak = Label()
        self.LabelWeak.grid(row=1,column=9)
        self.LabelMiddle = Label()
        self.LabelMiddle.grid(row=1,column=10)
        self.LabelPower = Label()
        self.LabelPower.grid(row=1,column=11)
        self.LabelHighSatisfy = Label()
        self.LabelHighSatisfy.grid(row=1,column=12)
        self.LabelClose = Label() 
        self.LabelClose.grid(row=3,column=10)
        
        self.test2 = StringVar()
        self.test1 = Label()
        self.test1.grid(row=10,column=10)
        self.FutureToday()
        
#############################################################

        for x in range(8,13,1):                         ###set the label name
            self.InTimeFutureLabel = Label()  
            self.InTimeFutureLabel["text"]=self.LabelName[x-8]
            self.InTimeFutureLabel.grid(row=0,column=x)

        self.ThreeMenDollar()
        self.PrintFeautre()
     #   self.InTimeObject.GetInTimeStockInfo(2330)
####################### initial interface #######################        
        self.interface.after(1000,self.PrintInTime)
        self.interface.mainloop()

#
    def mmka(self):
        while True:
            self.InTimeObject.FutureInTime()
            
            self.HighSatisfy.set(self.InTimeObject.PowerStage+self.InTimeObject.MiddleStage)
            self.showPowerStage.set(self.InTimeObject.WeakStage+self.InTimeObject.MiddleStage)
            self.showMiddleStage.set(self.InTimeObject.MiddleStage)
            self.showWeakStage.set(self.InTimeObject.MiddleStage-self.InTimeObject.WeakStage)       
            self.LowSatisfy.set(self.InTimeObject.MiddleStage-self.InTimeObject.PowerStage)
            self.showClosePrice.set(self.InTimeObject.FutureNowPrice)
          
            self.LabelHighSatisfy["text"]=self.HighSatisfy.get()
            self.LabelPower["text"] = self.showPowerStage.get()
            self.LabelMiddle["text"]=self.showMiddleStage.get()  
            self.LabelWeak["text"]=self.showWeakStage.get()
            self.LabelLowSatisfy["text"]=self.LowSatisfy.get()
            self.LabelClose["text"]=self.showClosePrice.get()
                    
            self.LabelHighSatisfy.update()
            self.LabelLowSatisfy.update()                             
            self.LabelMiddle.update()
            self.LabelPower.update()
            self.LabelWeak.update()
            self.LabelClose.update()
            time.sleep(1)
            
    def PrintInTime(self):
        self.PrintFeautre_thread = threading.Thread(target=self.mmka)
        self.PrintFeautre_thread.start()  
     
    def tesfile(self,event):
            count =0 
            f = open("export.csv","r")
            for row in csv.reader(f):
               # print(row)
                count+=1 
                if count >=4 and (row[2]!='-' and float(row[2])<=self.TargetPERatio):
                        QueryPrice=Share(row[0]+'.TW')
                      #  volumetemp =  self.InTimeObject.GetInTimeStockInfo(row[0])
                #             print(volumetemp)
                        if QueryPrice.get_earnings_share()!=None and float(QueryPrice.get_earnings_share())>=self.TargetEPS and float: #float(QueryPrice.get_earnings_share())>=self.TargetEPS:
                            if float(QueryPrice.get_volume())>=self.TargetVolume and float(QueryPrice.get_price())<=self.TargetPrice:
                                information = row[0] + row[1]+"\t" + QueryPrice.get_earnings_share() +" "+ QueryPrice.get_volume()+"   " + QueryPrice.get_price()
                                self.ThreemenBox.insert(END,information)
                if count==50:
                    break

         
    def PrintFeautre(self):
        temp=[]
        for item in csv.reader(codecs.iterdecode(self.TodayFutureResponse,"Latin-1")):
            temp.append(item[7])
        for x in range(temp.__len__()):
            if x!=0 and x<4:
                NewLabel=Label()
                NewLabel["text"]=temp[x]
                NewLabel.grid(row=x,column=5)
   #             print(temp[x])
        abc=[]
        blank=0
        #找尋三大法人  跳過地第一行csv
        for test in csv.reader(codecs.iterdecode(self.ThreeMenDollarResponse,"Latin-1")):
            if blank==0:            
                blank = blank+1
                continue
            abc.append(test[3])
        for y in range(abc.__len__()):
            if y >=2 and y <=4:
                NewLabel=Label()
                NewLabel["text"]=abc[y]
                NewLabel.grid(row=y-1,column=6)
        
        

    def GetTextFromVolumeField(self,event):
        self.TargetVolume = float(self.inputVolumeField.get())
        
    def GetTextFromWannaPriceField(self,event):
        self.TargetPrice = float(self.inputWannaPriceField.get())

    def GetTextFromPREField(self,event):
        self.TargetPERatio = float(self.inputPREField.get())

    def GetTextFromEPSField(self,event):
        self.TargetEPS=float(self.inputEPSField.get())
      
    def OutputData(self,event):
        for x in range(self.AllData.__len__()):     #Output the outcome 
            print(self.AllData[x])
        print(self.AllData.__len__())
        for x in range(int(self.AllData.__len__()/4)):
            for y in range(0,4,1):
               self.temp = tkinter.Text(self.interface,height=3,width=3)
               self.temp.insert(INSERT,self.AllData[x*4+y])
               self.temp.grid(row=5+x,column=y)


#b = GUI()

c = GetData()
