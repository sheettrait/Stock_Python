import openpyxl
import csv 
import datetime
from openpyxl import Workbook



wb = Workbook()
ws = wb.active
reader = csv.reader(open('Daily_2016_06_21.csv'))

row_counter = 1

for row in reader:
    try:
        if "TX" in row[1] and "MTX" not in row[1] and "201607" in row[2]:
            ws.cell(row = row_counter , column = 1, value = datetime.date(int(row[0][0:4]),int(row[0][4:6]),int(row[0][6:8])).isoformat())
            ws.cell(row = row_counter , column = 2, value = datetime.time(int(row[3][0:2]),int(row[3][2:4]),int(row[3][4:6])))
            ws.cell(row = row_counter , column = 3, value = row[4])
            row_counter = row_counter + 1
    except:
        pass

wb.save("DetailPrice.xlsx")